#!/usr/bin/env python3
"""
update_wc_fixtures.py — Replace FIXTURES array in WC dashboards from Excel source.

Reads input/WC/world-cup_2026.xlsx, extracts group stage matches (Matchday 1-3),
maps team names, and patches the const FIXTURES = [...] block in all 3 HTML files.
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).parent
INPUT_XLSX  = SCRIPT_DIR / 'world-cup_2026.xlsx'
HTML_FILES  = [
    SCRIPT_DIR / 'WorldCupDashboard-A.html',
    SCRIPT_DIR / 'WorldCupDashboard-B.html',
    SCRIPT_DIR / 'WorldCupDashboard-C.html',
]

# ── Team name normalisation (Excel → dashboard) ───────────────────────────────
TEAM_MAP = {
    'Czech Republic':          'Czechia',
    'Ivory Coast':             "Côte d\u2019Ivoire",
    'Bosnia & Herzegovina':    'Bosnia and Herzegovina',
    'Congo DR':                'DR Congo',
    'Cape Verde Islands':      'Cape Verde',
    'Türkiye':                 'Türkiye',   # already correct
}

# ── Group lookup built from dashboard TEAM_GROUPS ────────────────────────────
TEAM_GROUPS = {
    'Mexico':'A','South Korea':'A','South Africa':'A','Czechia':'A',
    'Bosnia and Herzegovina':'B','Canada':'B','Qatar':'B','Switzerland':'B',
    'Morocco':'C','Brazil':'C','Scotland':'C','Haiti':'C',
    'Australia':'D','Paraguay':'D','USA':'D','Türkiye':'D',
    'Germany':'E','Curaçao':'E',"Côte d\u2019Ivoire":'E','Ecuador':'E',
    'Netherlands':'F','Japan':'F','Sweden':'F','Tunisia':'F',
    'Belgium':'G','Egypt':'G','Iran':'G','New Zealand':'G',
    'Spain':'H','Cape Verde':'H','Saudi Arabia':'H','Uruguay':'H',
    'France':'I','Senegal':'I','Iraq':'I','Norway':'I',
    'Argentina':'J','Algeria':'J','Austria':'J','Jordan':'J',
    'Portugal':'K','DR Congo':'K','Uzbekistan':'K','Colombia':'K',
    'England':'L','Croatia':'L','Ghana':'L','Panama':'L',
}

def normalise(name):
    return TEAM_MAP.get(name, name)

def js_str(s):
    return s.replace('\\', '\\\\').replace("'", "\\'")

def load_fixtures():
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if not rec.get('Matchday'):
            continue  # skip KO rounds (TBD teams)
        rows.append(rec)

    fixtures = []
    for rec in rows:
        date    = str(rec['Date'])[:10]          # YYYY-MM-DD
        time    = str(rec['HKT Time'])            # HH:MM
        home    = normalise(str(rec['Home Team']))
        away    = normalise(str(rec['Away Team']))
        group   = TEAM_GROUPS.get(home) or TEAM_GROUPS.get(away)
        if not group:
            print(f"WARNING: no group found for {home} vs {away} — skipping")
            continue
        fixtures.append({'group': group, 'home': home, 'away': away, 'd': date, 't': time})

    # Sort by date then time
    fixtures.sort(key=lambda f: (f['d'], f['t']))
    return fixtures

def build_js_fixtures(fixtures):
    lines = ['const FIXTURES = [']
    for f in fixtures:
        h = js_str(f['home'])
        a = js_str(f['away'])
        lines.append(f"  {{group:'{f['group']}',home:'{h}',away:'{a}',d:'{f['d']}',t:'{f['t']}'}},")
    lines.append('];')
    return '\n'.join(lines)

def patch_html(path, new_fixtures_js):
    content = path.read_text(encoding='utf-8')
    pattern = r'const FIXTURES\s*=\s*\[[\s\S]*?\];'
    new_content, count = re.subn(pattern, new_fixtures_js, content, count=1)
    if count == 0:
        print(f"  WARNING: FIXTURES block not found in {path.name}")
        return False
    path.write_text(new_content, encoding='utf-8')
    print(f"  Patched {path.name}")
    return True

def main():
    print(f"Reading {INPUT_XLSX}")
    fixtures = load_fixtures()
    print(f"Loaded {len(fixtures)} group-stage fixtures")

    new_js = build_js_fixtures(fixtures)

    # Print preview of first 5 fixtures
    print("\nFirst 5 fixtures:")
    for f in fixtures[:5]:
        print(f"  {f['d']} {f['t']} HKT  Group {f['group']}: {f['home']} vs {f['away']}")

    for html_path in HTML_FILES:
        if not html_path.exists():
            print(f"  NOT FOUND: {html_path}")
            continue
        patch_html(html_path, new_js)

    print("\nDone. Commit and push to deploy.")

if __name__ == '__main__':
    main()
